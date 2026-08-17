"""
scripts/live_discovery_and_export.py

Helios v3.0 — Multi-Company Real Job Discovery & Structured Clean Export.
Scrapes 35+ top tech companies, extracts structured fields:
  - Company
  - Role / Title
  - Job Type (Intern / Fresher / PPO / Full-Time / Experienced)
  - Experience Level
  - Location / Remote
  - Salary / CTC / Stipend
  - Match Fit
  - Apply Link
"""
import os
import sys
import re
import csv
import json
import uuid
import asyncio
from datetime import datetime
from typing import Optional

base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if base_dir not in sys.path:
    sys.path.insert(0, base_dir)

from core.config.profile_loader import load_candidate_profile
from core.models.job import Job
from backend.src.connectors.ashby import AshbyConnector
from backend.src.connectors.greenhouse import GreenhouseConnector
from backend.src.connectors.lever import LeverConnector
from intelligence.embeddings.provider import LocalEmbeddingProvider
from intelligence.ranking.semantic_scorer import (
    SemanticScorer,
    generate_candidate_vector,
)
from intelligence.ranking.ranker import RankingAgent

# 35+ Top Tech Company Boards
ASHBY_COMPANIES = [
    "linear", "ramp", "openai", "vercel", "supabase", "notion", "sentry"
]

GREENHOUSE_COMPANIES = [
    "stripe", "figma", "airbnb", "postman", "razorpay", "browserstack",
    "databricks", "snowflake", "github", "cloudflare", "canva", "pinterest",
    "doordash", "robinhood", "discord", "plaid", "scaleai", "gusto", "mongodb"
]

LEVER_COMPANIES = [
    "cred", "spotify", "netflix", "inmobi", "meesho", "groww", "palantir"
]


def extract_job_type(title: str, description: str) -> str:
    t_lower = title.lower()
    d_lower = (description or "").lower()[:1000]

    if any(k in t_lower for k in ["intern", "internship", "trainee", "summer 202", "co-op", "apprentice"]):
        if "ppo" in t_lower or "ppo" in d_lower or "full-time" in t_lower:
            return "Internship + PPO"
        return "Internship"

    if any(k in t_lower for k in ["graduate", "new grad", "campus", "entry level", "fresher", "associate engineer", "sde 1", "sde i", "software engineer 1", "software engineer i", "junior"]):
        return "Fresher / Entry-Level (0-1 yrs)"

    if any(k in t_lower for k in ["senior", "lead", "staff", "principal", "director", "manager", "head"]):
        return "Senior / Lead (5+ yrs)"

    if any(k in t_lower for k in ["sde 2", "sde ii", "software engineer 2", "software engineer ii", "mid"]):
        return "Experienced (2-4 yrs)"

    if "0-1 year" in d_lower or "0-2 year" in d_lower or "fresher" in d_lower or "recent graduate" in d_lower:
        return "Fresher / Entry-Level (0-1 yrs)"
    if "3+ year" in d_lower or "5+ year" in d_lower:
        return "Experienced (3+ yrs)"

    return "Full-Time (General)"


def extract_experience(title: str, description: str, job_type: str) -> str:
    if "Intern" in job_type:
        return "0 yrs (College / Student)"
    if "Fresher" in job_type:
        return "0-1 yrs (Fresher / New Grad)"

    text = f"{title} {description or ''}"
    m = re.search(r"(\d+)\s*(?:-\s*(\d+))?\s*\+?\s*(?:years?|yrs?)(?:\s+of\s+experience)?", text, re.IGNORECASE)
    if m:
        low = m.group(1)
        high = m.group(2)
        if high:
            return f"{low}-{high} yrs"
        return f"{low}+ yrs"

    if "Senior" in job_type:
        return "5+ yrs"
    if "Experienced" in job_type:
        return "2-4 yrs"

    return "1-3 yrs (Standard)"


def extract_compensation(description: str, raw_data: dict) -> str:
    desc = description or ""

    if isinstance(raw_data, dict):
        comp = raw_data.get("compensation") or raw_data.get("salary") or raw_data.get("payRange")
        if comp:
            if isinstance(comp, dict):
                min_v = comp.get("min") or comp.get("minValue")
                max_v = comp.get("max") or comp.get("maxValue")
                curr = comp.get("currency") or "USD"
                if min_v and max_v:
                    return f"{curr} {min_v:,.0f} - {max_v:,.0f} / yr"
            elif isinstance(comp, str) and len(comp.strip()) > 3:
                return comp.strip()

    lpa_match = re.search(r"(?:₹|INR|Rs\.?)\s*(\d+(?:\.\d+)?)\s*(?:-\s*(\d+(?:\.\d+)?))?\s*(?:LPA|Lacs?|Lakhs?)", desc, re.IGNORECASE)
    if lpa_match:
        min_l = lpa_match.group(1)
        max_l = lpa_match.group(2)
        return f"₹{min_l}{(' - ' + max_l) if max_l else ''} LPA"

    usd_match = re.search(r"\$(\d{2,3}(?:,\d{3})+)\s*(?:-\s*\$(\d{2,3}(?:,\d{3})+))?", desc)
    if usd_match:
        min_s = usd_match.group(1)
        max_s = usd_match.group(2)
        return f"${min_s}{(' - $' + max_s) if max_s else ''} / yr"

    stipend_match = re.search(r"(\$|₹|INR)\s*(\d+[\d,]*)\s*(?:/hr|/hour|/month|per month|stipend)", desc, re.IGNORECASE)
    if stipend_match:
        curr = stipend_match.group(1)
        amt = stipend_match.group(2)
        return f"{curr}{amt} (Stipend / Rate)"

    return "Competitive (Market Standard)"


from backend.src.connectors.linkedin import LinkedInConnector


async def scrape_all_companies() -> list[Job]:
    all_jobs: list[Job] = []
    print(f"\n[Scraper] Querying {len(ASHBY_COMPANIES)} Ashby, {len(GREENHOUSE_COMPANIES)} Greenhouse, {len(LEVER_COMPANIES)} Lever companies, and LinkedIn Live...")

    # 1. Ashby Portals
    for comp in ASHBY_COMPANIES:
        try:
            c = AshbyConnector(site=comp)
            jobs = await c.search(query="", max_results=12)
            if jobs:
                print(f" -> Ashby [{comp.upper()}]: {len(jobs)} live jobs")
                all_jobs.extend(jobs)
        except Exception as e:
            pass

    # 2. Greenhouse Portals
    for comp in GREENHOUSE_COMPANIES:
        try:
            g = GreenhouseConnector(board_token=comp)
            jobs = await g.search(query="", max_results=12)
            if jobs:
                print(f" -> Greenhouse [{comp.upper()}]: {len(jobs)} live jobs")
                all_jobs.extend(jobs)
        except Exception as e:
            pass

    # 3. Lever Portals
    for comp in LEVER_COMPANIES:
        try:
            l = LeverConnector(site=comp)
            jobs = await l.search(query="", max_results=12)
            if jobs:
                print(f" -> Lever [{comp.upper()}]: {len(jobs)} live jobs")
                all_jobs.extend(jobs)
        except Exception as e:
            pass

    # 4. LinkedIn Live Search (Delhi / Noida / NCR / India & Major Companies)
    print("\n[Scraper] Querying LinkedIn for Delhi-NCR, Bangalore, and Top Companies (Bain, Flipkart, Amazon, Google, etc.)...")
    linkedin_conn = LinkedInConnector()

    india_queries = [
        ("Bain", "Delhi, India"),
        ("Flipkart", "Bengaluru, India"),
        ("Flipkart", "Delhi, India"),
        ("Swiggy", "Delhi, India"),
        ("Zomato", "Gurgaon, India"),
        ("Amazon", "Delhi, India"),
        ("Microsoft", "Noida, India"),
        ("Google", "Gurgaon, India"),
        ("Software Engineer", "Delhi, India"),
        ("Backend Developer", "Noida, India"),
        ("Python Developer", "Gurgaon, India"),
        ("Data Scientist", "Delhi, India"),
        ("Software Engineer Intern", "Delhi, India"),
        ("Software Engineer", "Bengaluru, India"),
    ]

    for q, loc in india_queries:
        try:
            jobs = await linkedin_conn.search(query=q, location=loc, max_results=6)
            if jobs:
                print(f" -> LinkedIn [{q} in {loc}]: {len(jobs)} live jobs")
                all_jobs.extend(jobs)
        except Exception as e:
            print(f" -> LinkedIn query error ({q} in {loc}): {e}")

    return all_jobs


def is_india_location(loc: str) -> bool:
    if not loc:
        return False
    l_lower = loc.lower()
    india_keywords = [
        "india", "delhi", "noida", "gurgaon", "gurugram", "ncr",
        "bangalore", "bengaluru", "mumbai", "hyderabad", "pune",
        "chennai", "kolkata", "ahmedabad", "jaipur"
    ]
    return any(k in l_lower for k in india_keywords)


async def run_live_benchmark():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    print("=" * 80)
    print(" HELIOS v3.0 — DUAL-TAB REAL JOB DISCOVERY (INDIA vs REMOTE/INTERNATIONAL)")
    print("=" * 80)

    # 1. Load Candidate Profile
    print("\n[Step 1] Loading Candidate Profile & generating 384-d embedding...")
    profile = load_candidate_profile()
    provider = LocalEmbeddingProvider()
    candidate_vec = await generate_candidate_vector(profile, provider)
    print(f" -> Candidate: {profile.name} ({profile.location})")

    scorer = SemanticScorer(candidate_vector=candidate_vec)
    ranker = RankingAgent(profile, semantic_scorer=scorer)

    # 2. Scrape Jobs
    print("\n[Step 2] Discovering live jobs across company career portals...")
    raw_jobs = await scrape_all_companies()
    print(f"\n[Total Discovered]: {len(raw_jobs)} live postings across ~35 top tech companies.")

    # 3. Embed, Rank, and Categorize into 2 Tabs
    print("\n[Step 3] Embedding and Ranking live jobs across 5 Dimensions...")
    india_rows = []
    remote_rows = []

    for job in raw_jobs:
        job_type = extract_job_type(job.title, job.description)
        exp_req = extract_experience(job.title, job.description, job_type)
        comp_str = extract_compensation(job.description, job.raw_data or {})

        text_to_embed = f"{job.title} at {job.company}. Location: {job.location}. {(job.description or '')[:300]}"
        job_vectors = await provider.embed([text_to_embed])
        job_vec = job_vectors[0] if job_vectors else []
        emb_id = str(uuid.uuid4())
        scorer.cache_vector(emb_id, job_vec)

        ranking = ranker.rank(job, embedding_id=emb_id)

        row = {
            "Company": job.company,
            "Role / Title": job.title,
            "Job Type": job_type,
            "Experience Level": exp_req,
            "Location / Remote": job.location or "Remote / Global",
            "Salary / CTC / Stipend": comp_str,
            "Match Fit": f"{int(ranking.overall_score * 100)}%",
            "Apply Link": job.apply_url or job.source_url,
        }

        if is_india_location(job.location):
            india_rows.append(row)
        else:
            remote_rows.append(row)

    # Sort each tab by match fit descending
    india_rows.sort(key=lambda r: int(r["Match Fit"].replace("%", "")), reverse=True)
    remote_rows.sort(key=lambda r: int(r["Match Fit"].replace("%", "")), reverse=True)

    print(f"\n[Separation Results]:")
    print(f" -> Tab 1: India (Delhi / Noida / NCR & Tech Hubs): {len(india_rows)} jobs")
    print(f" -> Tab 2: Remote & International (US / Europe / Global): {len(remote_rows)} jobs")

    # 4. Save CSVs
    os.makedirs(os.path.join(base_dir, "data"), exist_ok=True)
    fieldnames = [
        "Company",
        "Role / Title",
        "Job Type",
        "Experience Level",
        "Location / Remote",
        "Salary / CTC / Stipend",
        "Match Fit",
        "Apply Link",
    ]

    csv_india_path = os.path.join(base_dir, "data", "jobs_india_delhi_ncr.csv")
    csv_remote_path = os.path.join(base_dir, "data", "jobs_remote_international.csv")
    csv_master_path = os.path.join(base_dir, "data", "helios_live_jobs.csv")

    with open(csv_india_path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        w.writerows(india_rows)

    with open(csv_remote_path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        w.writerows(remote_rows)

    with open(csv_master_path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        w.writerows(india_rows + remote_rows)

    # 5. Save Multi-Tab Excel Workbook (.xlsx)
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "India (Delhi-NCR & Tech Hubs)"
    ws1.append(fieldnames)
    for r in india_rows:
        ws1.append([r[k] for k in fieldnames])

    ws2 = wb.create_sheet(title="Remote & International")
    ws2.append(fieldnames)
    for r in remote_rows:
        ws2.append([r[k] for k in fieldnames])

    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")

    for ws in [ws1, ws2]:
        for col_num, cell in enumerate(ws[1], 1):
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 26
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 14), 45)

    xlsx_path = os.path.join(base_dir, "data", "helios_jobs_two_tabs.xlsx")
    wb.save(xlsx_path)

    # 6. Save Dual-Tab Interactive HTML Dashboard
    html_content = generate_dual_tab_html_dashboard(india_rows, remote_rows, profile.name)
    html_path = os.path.join(base_dir, "data", "helios_live_jobs.html")
    with open(html_path, "w", encoding="utf-8") as f:
        f.write(html_content)

    print(f"\n[Export Completed]:")
    print(f" - Excel with 2 Tabs: {xlsx_path}")
    print(f" - Tab 1 CSV (India): {csv_india_path}")
    print(f" - Tab 2 CSV (Remote): {csv_remote_path}")
    print(f" - Dual-Tab HTML Dashboard: {html_path}")


def generate_dual_tab_html_dashboard(india_records: list[dict], remote_records: list[dict], candidate_name: str) -> str:
    def build_rows(records):
        out = []
        for r in records:
            score_val = int(r["Match Fit"].replace("%", ""))
            score_badge_class = "badge-green" if score_val >= 75 else ("badge-yellow" if score_val >= 60 else "badge-gray")
            out.append(f"""
            <tr>
                <td><span class="badge {score_badge_class}">{r['Match Fit']}</span></td>
                <td><strong>{r['Company']}</strong></td>
                <td>{r['Role / Title']}</td>
                <td><span class="type-tag">{r['Job Type']}</span></td>
                <td>{r['Experience Level']}</td>
                <td>{r['Location / Remote']}</td>
                <td><span class="salary-tag">{r['Salary / CTC / Stipend']}</span></td>
                <td><a href="{r['Apply Link']}" target="_blank" class="apply-btn">Apply Now &rarr;</a></td>
            </tr>
            """)
        return "".join(out)

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <title>Helios v3.0 — Discovered Jobs for {candidate_name}</title>
    <style>
        body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, Helvetica, Arial, sans-serif; background: #0f172a; color: #f8fafc; margin: 0; padding: 24px; }}
        .header {{ max-width: 1350px; margin: 0 auto 20px auto; display: flex; justify-content: space-between; align-items: center; }}
        h1 {{ margin: 0; font-size: 24px; color: #38bdf8; }}
        .stats {{ color: #94a3b8; font-size: 14px; }}
        .tabs {{ max-width: 1350px; margin: 0 auto 16px auto; display: flex; gap: 12px; }}
        .tab-btn {{ background: #1e293b; color: #94a3b8; border: 1px solid #334155; padding: 10px 20px; border-radius: 8px; font-weight: 600; font-size: 14px; cursor: pointer; transition: 0.2s; }}
        .tab-btn.active {{ background: #2563eb; color: #ffffff; border-color: #2563eb; }}
        .container {{ max-width: 1350px; margin: 0 auto; background: #1e293b; border-radius: 12px; overflow: hidden; box-shadow: 0 10px 25px rgba(0,0,0,0.5); }}
        table {{ width: 100%; border-collapse: collapse; text-align: left; font-size: 13px; }}
        th {{ background: #334155; padding: 14px 16px; color: #cbd5e1; font-weight: 600; text-transform: uppercase; font-size: 11px; letter-spacing: 0.5px; }}
        td {{ padding: 14px 16px; border-bottom: 1px solid #334155; }}
        tr:hover {{ background: #243248; }}
        .badge {{ padding: 4px 8px; border-radius: 6px; font-weight: 700; font-size: 12px; display: inline-block; }}
        .badge-green {{ background: #065f46; color: #34d399; }}
        .badge-yellow {{ background: #713f12; color: #facc15; }}
        .badge-gray {{ background: #374151; color: #9ca3af; }}
        .type-tag {{ background: #1e3a8a; color: #93c5fd; padding: 3px 6px; border-radius: 4px; font-size: 11px; font-weight: 600; }}
        .salary-tag {{ color: #a7f3d0; font-weight: 600; }}
        .apply-btn {{ background: #2563eb; color: #fff; text-decoration: none; padding: 6px 12px; border-radius: 6px; font-weight: 600; font-size: 12px; transition: 0.2s; display: inline-block; white-space: nowrap; }}
        .apply-btn:hover {{ background: #1d4ed8; }}
        .tab-content {{ display: none; }}
        .tab-content.active {{ display: block; }}
    </style>
</head>
<body>
    <div class="header">
        <div>
            <h1>Helios v3.0 — Live Discovered Jobs</h1>
            <div class="stats">Candidate: <strong>{candidate_name}</strong> | Total Discovered: <strong>{len(india_records) + len(remote_records)}</strong></div>
        </div>
    </div>
    <div class="tabs">
        <button class="tab-btn active" onclick="switchTab('india')">📍 India (Delhi-NCR & Tech Hubs) ({len(india_records)})</button>
        <button class="tab-btn" onclick="switchTab('remote')">🌐 Remote & International ({len(remote_records)})</button>
    </div>
    <div class="container">
        <div id="tab-india" class="tab-content active">
            <table>
                <thead>
                    <tr>
                        <th>Match</th>
                        <th>Company</th>
                        <th>Role / Title</th>
                        <th>Job Type</th>
                        <th>Experience</th>
                        <th>Location</th>
                        <th>Salary / CTC / Stipend</th>
                        <th>Action</th>
                    </tr>
                </thead>
                <tbody>
                    {build_rows(india_records)}
                </tbody>
            </table>
        </div>
        <div id="tab-remote" class="tab-content">
            <table>
                <thead>
                    <tr>
                        <th>Match</th>
                        <th>Company</th>
                        <th>Role / Title</th>
                        <th>Job Type</th>
                        <th>Experience</th>
                        <th>Location</th>
                        <th>Salary / CTC / Stipend</th>
                        <th>Action</th>
                    </tr>
                </thead>
                <tbody>
                    {build_rows(remote_records)}
                </tbody>
            </table>
        </div>
    </div>
    <script>
        function switchTab(tabId) {{
            document.querySelectorAll('.tab-btn').forEach(btn => btn.classList.remove('active'));
            document.querySelectorAll('.tab-content').forEach(c => c.classList.remove('active'));
            if (tabId === 'india') {{
                document.querySelectorAll('.tab-btn')[0].classList.add('active');
                document.getElementById('tab-india').classList.add('active');
            }} else {{
                document.querySelectorAll('.tab-btn')[1].classList.add('active');
                document.getElementById('tab-remote').classList.add('active');
            }}
        }}
    </script>
</body>
</html>
"""


if __name__ == "__main__":
    asyncio.run(run_live_benchmark())
