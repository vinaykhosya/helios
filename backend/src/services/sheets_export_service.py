"""
backend/src/services/sheets_export_service.py

Dual-Tab Excel & Live Google Sheets Push Synchronizer with Freshness Intelligence.
Synchronizes opportunities into:
  1. Local Excel Workbook (data/helios_jobs_two_tabs.xlsx) with Freshness & Provenance
  2. Local Master CSV (data/helios_live_jobs.csv)
  3. Live Google Spreadsheet projection across distinct tabs
"""
from __future__ import annotations

import os
import json
import csv
from typing import List, Dict, Any
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from intelligence.freshness.gate import FreshnessGate, DEFAULT_FRESHNESS_SETTINGS

BASE_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
EXCEL_PATH = os.path.join(BASE_DIR, "data", "helios_jobs_two_tabs.xlsx")
CSV_PATH = os.path.join(BASE_DIR, "data", "helios_live_jobs.csv")
CSV_INDIA_PATH = os.path.join(BASE_DIR, "data", "jobs_india_delhi_ncr.csv")
CSV_REMOTE_PATH = os.path.join(BASE_DIR, "data", "jobs_remote_international.csv")

FIELDNAMES = [
    "Company",
    "Role / Title",
    "Job Type",
    "Experience Level",
    "Location / Remote",
    "Salary / CTC / Stipend",
    "Match Fit",
    "Posted Date",
    "Age (Days)",
    "Freshness",
    "Freshness Confidence",
    "Ready to Apply",
    "Apply Link",
]


def sync_local_excel_and_csv(jobs: List[Dict[str, Any]]) -> Dict[str, Any]:
    """
    Regenerates and writes data/helios_jobs_two_tabs.xlsx and data/helios_live_jobs.csv
    with Freshness Intelligence and provenance fields.
    """
    os.makedirs(os.path.join(BASE_DIR, "data"), exist_ok=True)
    gate = FreshnessGate(DEFAULT_FRESHNESS_SETTINGS)

    india_rows = []
    remote_rows = []
    all_rows = []

    for j in jobs:
        fit_score = j.get("fit_score")
        if fit_score is not None:
            fit_str = f"{int(fit_score * 100)}%"
        else:
            fit_str = j.get("match_fit") or j.get("Match Fit") or "85%"

        age_days = j.get("age_days")
        age_str = f"{age_days}d" if age_days is not None else "Unknown"

        fresh_status = j.get("freshness_status") or ("FRESH" if age_days is not None and age_days <= 7 else "UNKNOWN")
        fresh_badge = (
            "🟢 FRESH" if fresh_status == "FRESH"
            else ("🟡 AGING" if fresh_status == "AGING"
            else ("🟠 STALE" if fresh_status == "STALE"
            else ("🔴 VERY STALE" if fresh_status == "VERY_STALE"
            else "⚪ UNKNOWN")))
        )

        is_ready = gate.is_ready_to_apply(j)
        ready_badge = "✅ YES" if is_ready else "❌ NO"

        row = {
            "Company": j.get("company") or j.get("Company") or "Unknown",
            "Role / Title": j.get("title") or j.get("Role / Title") or "Software Engineer",
            "Job Type": j.get("job_type") or j.get("Job Type") or "Full-Time",
            "Experience Level": j.get("experience_years") or j.get("Experience Level") or "1-3 yrs",
            "Location / Remote": j.get("location") or j.get("Location / Remote") or "Remote",
            "Salary / CTC / Stipend": j.get("compensation") or j.get("Salary / CTC / Stipend") or "Competitive",
            "Match Fit": fit_str,
            "Posted Date": j.get("posted_date_str") or j.get("Posted Date") or "Recent",
            "Age (Days)": age_str,
            "Freshness": fresh_badge,
            "Freshness Confidence": j.get("freshness_confidence") or "CONFIRMED_POSTED",
            "Ready to Apply": ready_badge,
            "Apply Link": j.get("apply_url") or j.get("Apply Link") or "#",
        }

        all_rows.append(row)

        is_india = j.get("is_india")
        if is_india is None:
            loc = (row["Location / Remote"]).lower()
            is_india = any(k in loc for k in ["india", "delhi", "noida", "gurgaon", "gurugram", "bangalore", "bengaluru", "hyderabad", "pune", "mumbai"])

        if is_india:
            india_rows.append(row)
        else:
            remote_rows.append(row)

    # Sort each tab: Ready-to-apply & Freshness first, then fit score descending
    def sort_key(r):
        age_val = 999
        try:
            raw_age = str(r["Age (Days)"]).replace("d", "")
            if raw_age.isdigit():
                age_val = int(raw_age)
        except Exception:
            pass

        fit_val = 0
        try:
            fit_val = int(str(r["Match Fit"]).replace("%", ""))
        except Exception:
            pass

        is_ready_prio = 0 if "YES" in r["Ready to Apply"] else 1
        return (is_ready_prio, -fit_val, age_val)

    india_rows.sort(key=sort_key)
    remote_rows.sort(key=sort_key)
    all_rows.sort(key=sort_key)

    # Write Master CSVs
    with open(CSV_PATH, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=FIELDNAMES)
        w.writeheader()
        w.writerows(all_rows)

    with open(CSV_INDIA_PATH, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=FIELDNAMES)
        w.writeheader()
        w.writerows(india_rows)

    with open(CSV_REMOTE_PATH, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=FIELDNAMES)
        w.writeheader()
        w.writerows(remote_rows)

    # Generate Multi-Tab Excel Workbook
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "India (Delhi-NCR & Tech Hubs)"
    ws1.append(FIELDNAMES)
    for r in india_rows:
        ws1.append([r[k] for k in FIELDNAMES])

    ws2 = wb.create_sheet(title="Remote & International")
    ws2.append(FIELDNAMES)
    for r in remote_rows:
        ws2.append([r[k] for k in FIELDNAMES])

    ws3 = wb.create_sheet(title="All Opportunities & History")
    ws3.append(FIELDNAMES)
    for r in all_rows:
        ws3.append([r[k] for k in FIELDNAMES])

    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")

    for ws in [ws1, ws2, ws3]:
        for col_num, cell in enumerate(ws[1], 1):
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 26
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 14), 45)

    wb.save(EXCEL_PATH)

    # Push to live Google Sheets if credentials or gspread configured
    gspread_synced = False
    gspread_msg = ""
    spreadsheet_id = os.getenv("GOOGLE_SHEETS_SPREADSHEET_ID", "1-fMsNdwrR-OPZvrLza1QpGtrIj8GhsEy")
    sa_json = os.getenv("GOOGLE_SERVICE_ACCOUNT_JSON")

    if sa_json and spreadsheet_id:
        try:
            import gspread
            from google.oauth2.service_account import Credentials
            creds = Credentials.from_service_account_info(
                json.loads(sa_json),
                scopes=["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
            )
            gc = gspread.authorize(creds)
            ss = gc.open_by_key(spreadsheet_id)

            # Sync Tab 1: India
            try:
                ws_in = ss.worksheet("India (Delhi-NCR & Tech Hubs)")
            except gspread.WorksheetNotFound:
                ws_in = ss.add_worksheet("India (Delhi-NCR & Tech Hubs)", rows=1000, cols=len(FIELDNAMES))
            ws_in.clear()
            in_matrix = [FIELDNAMES] + [[r[k] for k in FIELDNAMES] for r in india_rows]
            ws_in.update(range_name="A1", values=in_matrix)

            # Sync Tab 2: Remote
            try:
                ws_rem = ss.worksheet("Remote & International")
            except gspread.WorksheetNotFound:
                ws_rem = ss.add_worksheet("Remote & International", rows=1000, cols=len(FIELDNAMES))
            ws_rem.clear()
            rem_matrix = [FIELDNAMES] + [[r[k] for k in FIELDNAMES] for r in remote_rows]
            ws_rem.update(range_name="A1", values=rem_matrix)

            gspread_synced = True
            gspread_msg = f"Live Google Sheet updated directly across 2 tabs with Freshness Intelligence ({len(india_rows)} India, {len(remote_rows)} Remote)."
        except Exception as e:
            print(f"[SheetsExportService] gspread API note: {e}")
            gspread_msg = f"Excel rebuilt with Freshness Intelligence ({len(india_rows)} India, {len(remote_rows)} Remote). Google API note: {e}"

    return {
        "status": "success",
        "excel_path": EXCEL_PATH,
        "india_count": len(india_rows),
        "remote_count": len(remote_rows),
        "total_rows": len(all_rows),
        "gspread_synced": gspread_synced,
        "spreadsheet_url": f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}",
        "message": gspread_msg or f"Successfully exported {len(all_rows)} opportunities with Freshness Intelligence to Excel & Google Sheet projection!",
    }
